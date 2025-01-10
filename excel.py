from openpyxl import load_workbook
import json
from typing import Dict, List, Any, Optional
from pathlib import Path
import io
from datetime import datetime

class ExcelDocumentParser:
    def __init__(self, file_content: bytes, file_name: str = ""):
        """Excel 문서 파서 초기화"""
        self.wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        # 보이는 시트만 필터링
        self.visible_sheets = []
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            if sheet.sheet_state == 'visible':
                self.visible_sheets.append(sheet_name)
        
        # 기본 메타데이터 구성
        self.document_structure = {
            'metadata': {
                'file_name': file_name,
                'total_sheets': len(self.visible_sheets),
                'sheet_names': self.visible_sheets,
                'sheets_info': {},
                'last_modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            },
            'sheets': {}
        }

    def _get_sheet_info(self, sheet) -> Dict:
        """시트의 기본 정보 추출"""
        return {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'has_merged_cells': bool(sheet.merged_cells),
            'sheet_state': sheet.sheet_state  # 시트 상태 추가
        }

    def _get_cell_value(self, cell: Any) -> Optional[Any]:
        """셀 값을 적절한 형태로 반환"""
        if cell.value is None:
            return None
        
        value = cell.value
        try:
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d')
            return value
        except:
            return str(value)

    def parse_sheet(self, sheet) -> Dict:
        """시트 내용 파싱"""
        sheet_content = []
        current_row_index = 0

        for row in sheet.rows:
            row_content = {}
            has_value = False

            for cell in row:
                value = self._get_cell_value(cell)
                if value is not None:  # 빈 셀 제외
                    has_value = True
                    row_content[cell.column_letter] = {
                        'value': value,
                        'coordinate': cell.coordinate
                    }
            
            if has_value:  # 값이 있는 행만 추가
                sheet_content.append({
                    'row_index': current_row_index,
                    'content': row_content
                })
            
            current_row_index += 1

        return sheet_content

    def parse_document(self) -> Dict:
        """전체 문서 파싱 (숨겨진 시트 제외)"""
        for sheet_name in self.visible_sheets:
            sheet = self.wb[sheet_name]
            
            # 시트 정보 저장
            self.document_structure['metadata']['sheets_info'][sheet_name] = self._get_sheet_info(sheet)
            
            # 시트 내용 파싱
            sheet_content = self.parse_sheet(sheet)
            if sheet_content:  # 내용이 있는 시트만 추가
                self.document_structure['sheets'][sheet_name] = sheet_content

        return self.document_structure

def process_excel_content(file_content: bytes, file_name: str = "") -> Dict:
    """Excel 문서를 처리하고 구조화된 형태로 반환"""
    try:
        parser = ExcelDocumentParser(file_content, file_name)
        return parser.parse_document()
            
    except Exception as e:
        print(f'파일 처리 중 오류 발생: {str(e)}')
        raise